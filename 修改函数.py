import openpyxl
import re

# 打开 Excel 文件
workbook = openpyxl.load_workbook(r'C:\Users\22327\Desktop\3.xlsx')

# 选择要操作的工作表
sheet = workbook['Sheet16']

# 选择要操作的列（第I列，列号9）
column_i = 9  # 列号从1开始，所以第I列对应列号9
column_g = 8  # 假设要从第G列（列号7）获取相应的小数

# 遍历第I列，查找包含 COUNTIF 函数的公式，将 "<15" 替换为 `<` 加上第G列中对应单元格中的小数
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=column_i, max_col=column_i):
    for cell in row:
        if cell.data_type == 'f' and 'COUNTIF(' in cell.value:  # 检查单元格是否包含 COUNTIF 函数
            # 从第G列中获取对应的小数
            corresponding_cell = sheet.cell(row=cell.row, column=column_g)
            corresponding_value = corresponding_cell.value

            # 如果对应单元格的值可以转换为浮点数，则进行替换
            if isinstance(corresponding_value, (int, float)):
                formula = cell.value
                # 使用正则表达式替换 "<15"
                formula = re.sub(r'<14.5', f'<{corresponding_value}', formula)
                cell.value = formula

# 保存更改
workbook.save('33.xlsx')
