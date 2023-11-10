import openpyxl

# 打开 Excel 文件
workbook = openpyxl.load_workbook(r'C:\Users\22327\Desktop\3.xlsx')

# 选择要操作的工作表
sheet = workbook['Sheet16']

# 创建一个 Formula 对象，包含要粘贴的函数
specific_formula = '=COUNTIF(E2:E11592,"<15")-COUNTIF(E2:E11592,"<14.5")'

# 选择要操作的列（第I列，列号9）
column = 9  # 列号从1开始，所以第I列对应列号9

# 遍历指定列，将特定函数粘贴到每个单元格
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=column, max_col=column):
    for cell in row:
        cell.value = specific_formula

# 保存更改
workbook.save('33.xlsx')
